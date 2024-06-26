import time
from bs4 import BeautifulSoup
import requests
from selenium import webdriver
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re
from selenium_stealth import stealth

PATH = 'D:/nouran/programs/Chrome driver/chromedriver.exe'   # path to chrome driver

options = webdriver.ChromeOptions()
options.add_argument('start-maximize')
options.add_argument('--headless')
options.add_experimental_option('excludeSwitches',['enable-automation'])
options.add_experimental_option('useAutomationExtension',False)

driver = webdriver.Chrome(PATH, options=options)
stealth(driver,
        languages=['en=US','en'],
        vendor='Google Inc.',
        platform='Win32',
        webgl_vendor='Intel Inc.',
        renderer='Intel Iris OpenGL Engine',
        fix_hairline=True,)



products_count = 1   # 1 for first row in excel is the headers

def findURL(product):
    product_link = product.find('a', class_ = 'product-item-link')
    if product_link != None:
        product_link = product_link['href']
    else:
        product_link = 'Not Found'
        ws_errors.append(['product link',brand_name,products_count])

    return product_link

def findModelNameFromDiv(product_mainpage):
    model_name_div = product_mainpage.find('div', class_ = 'product attribute description')
    if model_name_div == None:
        model_name = ' '
    else:
        model_nameh1 = model_name_div.find('h1')
        if model_nameh1 != None:
            model_namestrong = model_nameh1.strong
            if model_namestrong == None:
                model_name = model_nameh1.text.strip()
            else:
                model_name = model_namestrong.text.strip()  
        else:
            model_name = ' '
    
    return model_name


def findModelName(product_mainpage,product_link):
    specs_table = product_mainpage.find('table', id = 'product-attribute-specs-table')
    if specs_table == None:
        model_name = findModelNameFromDiv(product_mainpage)

    else:
        model_name = specs_table.find('td', attrs={"data-th": "Model"})
        if model_name == None:
            model_name = findModelNameFromDiv(product_mainpage)

        else:
            model_name = model_name.text.strip()

    if model_name.strip() == '':
        ws_errors.append(['model name', product_link])

    return model_name



def findDescription(product,product_link):
    info_tag = product.find('a', class_ = 'product-item-link')
    if info_tag == None:
        ws_errors.append(['description',product_link])
        description = ' '
    else:
        description = info_tag.text.strip()
    
    return description

def findMoreDetails(product_mainpage, product_link):
    specs_table = product_mainpage.find('table', id = 'product-attribute-specs-table')
    if specs_table == None:
        details = product_mainpage.find('div', class_ = 'product attribute overview')
        if details == None:
            details = 'Not found'
            ws_errors.append(['details', product_link])
        else:
            details = details.text
    else:
        specs_details = specs_table.text.strip().splitlines(True)
        details = "".join([s for s in specs_details if s.strip()])
    
    return details


def findPrice(product,product_link):
    price = product.find('span', class_ = 'special-price')
    if price == None:
        price = product.find('span', class_ = 'price-wrapper').span.text.strip()
        if price == None:
            ws_errors.append(['price',product_link])
            price = ' '
    else:
        price = price.find('span', class_ = 'price')
        if price == None:
            price = ' '
            ws_errors.append(['price2',product_link])
        else:
            price = price.text.strip()

    return price


def findRating(product,product_link):
    rating = product.find('div', class_ = 'rating-result')
    if rating == None:
        ws_errors.append(['rating',product_link])
        rating = ' '
    else:
        rating = rating['title']
    
    return rating


def findStock(product_mainpage,product_link):
    availability = product_mainpage.find('div', title = 'Availability') 
    if availability == None:
        ws_errors.append(['stock1',product_link])
        stock = ' '
    else:
        availability = availability['class']
        if 'available' in availability:
            stock = product_mainpage.find('div', class_ = 'availability only')
            if stock == None:
                stock = 'In Stock'
            else:
                stock = stock['title']
        else:
            stock = 'Out of Stock'

    return stock


def findImage(product,product_link):
    image_link = product.find('img')
    if image_link == None:
        ws_errors.append(['image',product_link])
        image_link = 'Not Found'
    else:
        image_link = image_link['data-src']

    return image_link


def findReviews(product_link, product_name):
    url = product_link + '#reviews'
    driver.get(url)
    time.sleep(5)
    product_mainpage = BeautifulSoup(driver.page_source, 'html.parser')

    reviews_container = product_mainpage.find('ol', class_ = 'items review-items')
    if reviews_container == None:
        pass
    else:
        reviews_list = reviews_container.find_all('li', class_ = 'item review-item')

        if reviews_list == None:
            ws_errors.append(['review li',product_link])
        else:
            for review in reviews_list:
                title = review.find('div', class_ = 'review-title').text.strip()
                detail = review.find('div', class_ = 'review-content').text.strip()
                user = review.find('p', class_ = 'review-author').text.strip()
                date = review.find('p', class_ = 'review-date').text.strip()
                date = (date[10:])[:5] + '20' + (date[10:])[5:]
                ratings= review.find_all('div', class_ = 'rating-result')
            
                for i in [0,1,2]:
                    ratings[i] = ratings[i]['title']

                ws_reviews.append([product_name,user[3:],date,ratings[0][:-1],ratings[1][:-1],ratings[2][:-1],title,detail, product_link])
                



def findProducts(products_list):
    global products_count
    for product in products_list:
        products_count += 1
        print(products_count)

        if products_count % 15 == 0:
            time.sleep(60)

        product_link = findURL(product)
        description = findDescription(product,product_link)
        price = findPrice(product,product_link)
        rating = findRating(product,product_link)
        image_link = findImage(product,product_link)

        product_page = requests.get(product_link).text                               # product from inside
        product_mainpage = BeautifulSoup(product_page, 'html.parser')

        more_details = findMoreDetails(product_mainpage,product_link)
        model_name = findModelName(product_mainpage,product_link)
        if model_name.strip() == '':
            model_name = (re.split(r"[-,)(/]",description, maxsplit=1, flags=re.IGNORECASE))[0].strip()

        stock = findStock(product_mainpage,product_link)
        findReviews(product_link, model_name)

        try:
            ws_products.append([model_name,category_name,subcategory_name,brand_name,description,more_details,price,rating,stock])
        except:
            more_details =  ILLEGAL_CHARACTERS_RE.sub(r'',more_details)
            ws_products.append([model_name,category_name,subcategory_name,brand_name,description,more_details,price,rating,stock])

        if image_link == 'Not Found':
            ws_products.cell(row=products_count, column=10).value = image_link
        else:
            ws_products.cell(row=products_count, column=10).hyperlink = image_link
            ws_products.cell(row=products_count, column=10).value = "image"
            ws_products.cell(row=products_count, column=10).style = "Hyperlink"

        if product_link == 'Not Found':
            ws_products.cell(row=products_count, column=11).value = product_link
        else:
            ws_products.cell(row=products_count, column=11).hyperlink = product_link
            ws_products.cell(row=products_count, column=11).value = "URL"
            ws_products.cell(row=products_count, column=11).style = "Hyperlink"

        wb.save('2B Data.xlsx')


wb = Workbook()
ws_categories = wb.active
ws_categories.title = 'Categories'
ws_categories.append(['Category','SubCategory','Brand'])

ws_products = wb.create_sheet("Products")
ws_products.append(['Product Name','Category','Subcategory','Brand','Description','More details','Price','Rating','Stock','Image','URL'])

ws_reviews = wb.create_sheet("Reviews")
ws_reviews.append(['Product Name','User Name','Date','Quality Rating','Price Rating','Value Rating','Review Title','Review Content','Product Link'])

ws_errors = wb.create_sheet("Errors")
ws_errors.append(['Error Field','Product URL'])

driver = webdriver.Chrome(PATH)
driver.get('https://2b.com.eg/en/')

main_page = driver.page_source

#driver.quit()
webpage_home = BeautifulSoup(main_page, 'html.parser')

categories = webpage_home.find_all('li', class_ = 'ui-menu-item level0 fullwidth parent')

for category in categories:
    category_name = category.find('span').text.strip()                                   # category
    print(category_name)
    SubCategories = category.find_all('li', class_ = 'ui-menu-item level1 parent')
    SubCategories += category.find_all('li', class_ = 'ui-menu-item level1')

    for SubCategory in SubCategories:
        subcategory_name = SubCategory.find('span').text.strip()                       # subcategory
        print(subcategory_name)
        Brands = SubCategory.find_all('li', class_ = 'ui-menu-item level2')

        if Brands == []:
            brand_name = ' '
            ws_categories.append([category_name,subcategory_name,brand_name])
            products_page_link = SubCategory.find('a')['href']
            products_page = requests.get(products_page_link).text
            products_page = BeautifulSoup(products_page, 'html.parser')
            products = products_page.find('ol', class_ = 'filterproducts products list items product-items')
            if products == None:
                ws_errors.append(['products ol',brand_name])
            else:
                products_list = products.find_all('li')
                if products_list == []:
                    ws_errors.append(['products list li',brand_name])
                else:
                    findProducts(products_list)


        else:
            for brand in Brands:
                brand_name = brand.find('span').text.strip() 
                print(brand_name)                       # brand
                ws_categories.append([category_name,subcategory_name,brand_name])

                brand_page_link = brand.a['href']

                brand_products_page = requests.get(brand_page_link).text
                brand_page = BeautifulSoup(brand_products_page, 'html.parser')

                products = brand_page.find('ol', class_ = 'filterproducts products list items product-items')
                if products == None:
                    ws_errors.append(['products ol',brand_name])
                else:
                    products_list = products.find_all('li')
                    if products_list == []:
                        ws_errors.append(['products list li',brand_name])
                    else:
                        findProducts(products_list)








